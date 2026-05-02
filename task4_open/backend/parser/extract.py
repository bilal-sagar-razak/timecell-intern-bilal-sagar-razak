"""File-format extraction. Returns ExtractedContent for the LLM normalizer."""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Literal

ExtractKind = Literal["tables", "text", "image"]


class ExtractError(Exception):
    """Raised when a file cannot be read (corrupt / unsupported)."""


@dataclass
class ExtractedContent:
    """What extract() returns to normalize()."""
    kind: ExtractKind
    tables: list[tuple[str, list[list[str]]]] = field(default_factory=list)
    text: str = ""
    image_bytes: bytes = b""
    image_media_type: str = ""
    format_hint: str = "unknown"


def _cells_to_str(rows: list[list[object]]) -> list[list[str]]:
    """Stringify cells; None → empty string."""
    return [["" if c is None else str(c) for c in r] for r in rows]


def extract_xlsx(path: Path) -> ExtractedContent:
    """openpyxl extractor. For multi-sheet files, prefers a 'Combined' sheet if present."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames
    preferred = next((s for s in sheet_names if s.lower() == "combined"), None)
    if preferred:
        sheet_names = [preferred]

    tables = []
    for sn in sheet_names:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        while rows and all(c is None for c in rows[-1]):
            rows.pop()
        tables.append((sn, _cells_to_str(rows)))

    fmt = "zerodha_console" if "Combined" in wb.sheetnames or "Mutual Funds" in wb.sheetnames else "groww_xlsx"
    return ExtractedContent(kind="tables", tables=tables, format_hint=fmt)


def extract_xls(path: Path) -> ExtractedContent:
    """Handles both real binary .xls and HTML-disguised .xls (Camsonline format)."""
    head = path.read_bytes()[:200]
    if b"<html" in head.lower() or b"<!doctype html" in head.lower():
        text = path.read_text(encoding="utf-8", errors="replace")
        return ExtractedContent(kind="text", text=text, format_hint="camsonline_mhtml")
    try:
        import xlrd
        wb = xlrd.open_workbook(path)
        tables = []
        for sn in wb.sheet_names():
            ws = wb.sheet_by_name(sn)
            rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
            tables.append((sn, _cells_to_str(rows)))
        return ExtractedContent(kind="tables", tables=tables, format_hint="binary_xls")
    except Exception as e:
        raise ExtractError(f"could not read .xls file: {e}") from e


def extract_pdf(path: Path) -> ExtractedContent:
    """pdfplumber: page-by-page text extraction."""
    import pdfplumber
    parts = []
    try:
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                t = page.extract_text() or ""
                parts.append(f"--- page {i} ---\n{t}")
    except Exception as e:
        raise ExtractError(f"could not read PDF: {e}") from e
    return ExtractedContent(kind="text", text="\n\n".join(parts), format_hint="pdf")


def extract_image(path: Path) -> ExtractedContent:
    """Image: return bytes + media type for vision-capable LLM call."""
    suffix = path.suffix.lower()
    media = "image/png" if suffix == ".png" else "image/jpeg"
    return ExtractedContent(
        kind="image",
        image_bytes=path.read_bytes(),
        image_media_type=media,
        format_hint="image",
    )


def extract(path: Path) -> ExtractedContent:
    """Dispatch on file extension. Raises ExtractError on unsupported / unreadable."""
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return extract_xlsx(path)
    if suffix == ".xls":
        return extract_xls(path)
    if suffix == ".pdf":
        return extract_pdf(path)
    if suffix in (".png", ".jpg", ".jpeg"):
        return extract_image(path)
    raise ExtractError(f"unsupported file extension: {suffix}")
