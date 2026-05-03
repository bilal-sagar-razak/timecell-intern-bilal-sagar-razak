"""HDFC AMC monthly portfolio-disclosure adapter (xlsx)."""
from __future__ import annotations

import logging
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from amfi.schema import FundHolding, Scheme

logger = logging.getLogger(__name__)
AMC_NAME = "HDFC"

_NAME_COL_HEADERS = {"name of the instrument", "name of instrument", "instrument name"}
_ISIN_COL_HEADERS = {"isin"}
_PCT_COL_HEADERS = {"% to nav", "% of nav", "%nav"}
_VALUE_COL_HEADERS = {"market value (rs. in lakhs)", "market value (rs in lakhs)", "market value"}


def _find_col(headers: list, candidates: set[str]) -> int | None:
    for i, h in enumerate(headers):
        if h is None:
            continue
        s = str(h).strip().lower()
        if s in candidates:
            return i
    return None


def parse(file_path: Path) -> list[Scheme]:
    """Parse one HDFC monthly disclosure xlsx into Scheme objects (one per worksheet)."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    schemes: list[Scheme] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c).strip() if c else None for c in rows[0]]
        name_idx = _find_col(headers, _NAME_COL_HEADERS)
        isin_idx = _find_col(headers, _ISIN_COL_HEADERS)
        pct_idx = _find_col(headers, _PCT_COL_HEADERS)
        val_idx = _find_col(headers, _VALUE_COL_HEADERS)
        if name_idx is None or pct_idx is None:
            logger.warning("[hdfc] sheet %s has no recognized headers, skipping", ws.title)
            continue
        holdings: list[FundHolding] = []
        cash_pct = 0.0
        for row in rows[1:]:
            if not row or row[name_idx] is None:
                continue
            name = str(row[name_idx]).strip()
            pct_raw = row[pct_idx]
            if pct_raw is None:
                continue
            try:
                pct = float(pct_raw)
            except (TypeError, ValueError):
                continue
            isin = str(row[isin_idx]).strip() if (isin_idx is not None and row[isin_idx]) else None
            value = float(row[val_idx]) * 100000 if (val_idx is not None and row[val_idx] is not None) else 0.0
            if "cash" in name.lower() or (isin is None and "equiv" in name.lower()):
                cash_pct = pct
                continue
            holdings.append(FundHolding(
                name=name, isin=isin, weight_pct=pct, value_inr=value, kind="equity",
            ))
        if not holdings:
            continue
        schemes.append(Scheme(
            scheme_name=ws.title,
            isin=None,
            amc=AMC_NAME,
            scheme_aum_inr=sum(h.value_inr for h in holdings) / max(0.01, sum(h.weight_pct for h in holdings) / 100),
            as_of_date=date.today(),
            holdings=holdings,
            cash_pct=cash_pct,
        ))
    return schemes
