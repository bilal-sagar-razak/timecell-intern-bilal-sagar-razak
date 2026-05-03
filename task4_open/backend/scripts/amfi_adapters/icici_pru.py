"""ICICI Pru AMC monthly portfolio-disclosure adapter (xlsx)."""
from __future__ import annotations

import logging
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from amfi.schema import FundHolding, Scheme

logger = logging.getLogger(__name__)
AMC_NAME = "ICICI Pru"

_NAME_HEADERS = {"instrument", "name of the instrument", "name of instrument"}
_ISIN_HEADERS = {"isin"}
_PCT_HEADERS = {"% to net assets", "% to nav", "% nav"}
_VALUE_HEADERS = {"market value (rs lakhs)", "market value (rs. in lakhs)", "market value"}
_DEBT_HINTS = ("goi", "g-sec", "ncd", "bond", "debenture", "treasury", "treps", "g sec")


def _find(headers: list, candidates: set[str]) -> int | None:
    for i, h in enumerate(headers):
        if h is None:
            continue
        if str(h).strip().lower() in candidates:
            return i
    return None


def _kind_for(name: str) -> str:
    n = name.lower()
    if any(h in n for h in _DEBT_HINTS):
        return "debt"
    return "equity"


def parse(file_path: Path) -> list[Scheme]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    schemes: list[Scheme] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c).strip() if c else None for c in rows[0]]
        name_i = _find(headers, _NAME_HEADERS)
        isin_i = _find(headers, _ISIN_HEADERS)
        pct_i = _find(headers, _PCT_HEADERS)
        val_i = _find(headers, _VALUE_HEADERS)
        if name_i is None or pct_i is None:
            logger.warning("[icici_pru] sheet %s missing headers, skipping", ws.title)
            continue
        holdings: list[FundHolding] = []
        cash_pct = 0.0
        for row in rows[1:]:
            if not row or row[name_i] is None:
                continue
            name = str(row[name_i]).strip()
            pct_raw = row[pct_i]
            if pct_raw is None:
                continue
            try:
                pct = float(pct_raw)
            except (TypeError, ValueError):
                continue
            isin = str(row[isin_i]).strip() if (isin_i is not None and row[isin_i]) else None
            value = float(row[val_i]) * 100000 if (val_i is not None and row[val_i] is not None) else 0.0
            n_lower = name.lower()
            if "cash" in n_lower or "treps" in n_lower:
                cash_pct = pct
                continue
            holdings.append(FundHolding(
                name=name, isin=isin, weight_pct=pct, value_inr=value, kind=_kind_for(name),
            ))
        if not holdings:
            continue
        schemes.append(Scheme(
            scheme_name=ws.title,
            isin=None,
            amc=AMC_NAME,
            scheme_aum_inr=sum(h.value_inr for h in holdings) / max(0.01, sum(h.weight_pct for h in holdings) / 100),
            as_of_date=date.today(),
            holdings=holdings,
            cash_pct=cash_pct,
        ))
    return schemes
